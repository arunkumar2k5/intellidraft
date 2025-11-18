import json
import re
import openpyxl
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

# Error code table from the image
ERROR_TABLE = {
    '0': 'HS_NO_ERROR',
    '1': 'HS_AUDIO_SPI_ERROR',
    '2': 'HS_AUDIO_THREAD_LOCK_ERROR',
    '3': 'HS_AUDIO_THREAD_EXIT_ERROR',
    '4': 'HS_ETCO2_CONFIG_NOT_SET_ERROR',
    '5': 'HS_ETCO2_INIT_ERROR',
    '6': 'HS_ETCO2_THREAD_EXIT_ERROR',
    '7': 'HS_SPO2_INIT_ERROR',
    '8': 'HS_SPO2_CONN_ERROR',
    '9': 'HS_SPO2_THREAD_EXIT_ERROR',
    'A': 'HS_NBP_INIT_ERROR',
    'B': 'HS_NBP_CONN_ERROR',
    'C': 'HS_NBP_THREAD_EXIT_ERROR',
    'D': 'HS_ADAS_SPORT_ERROR',
    'E': 'HS_ADAS_THREAD_EXIT_ERROR',
    'F': 'HS_ECG_CORE_ERROR',
    '10': 'HS_ECG_DISCONNECT_ERROR',
    '11': 'HS_GPA_DISCONNECT_ERROR',
    '12': 'HS_WDOG_TIMEOUT_ERROR',
    '13': 'HS_FW_UPDATE_ERROR',
    '14': 'HS_FW_WRITE_FAIL',
    '15': 'HS_TCP_THREAD_EXIT_ERROR',
    '16': 'HS_GPA_CORE_ERROR',
    '17': 'HS_SPO2_FRAME_ERROR',
    '18': 'HS_SPO2_PARITY_ERROR',
    '19': 'HS_NBP_PARITY_ERROR',
    '1A': 'HS_NBP_FRAME_ERROR',
    '1B': 'HS_NBP_OVERRUN_ERROR',
    '1C': 'HS_NBP_OVERRUN_ERROR',
    '1D': 'HS_ADAS_ECG_BUFFER_OVERRUN',
    '1E': 'HS_ADAS_GPA_BUFFER_OVERRUN',
}


def append_events_to_excel(hisi_disconnects, client_shutdowns, hisib_errors, excel_path):
    # Sheet 1: Hisi Disconnects
    hd_headers = ["Date", "Time", "Last sample", "First sample"]
    # Sheet 2: Client Shutdowns
    cs_headers = ["Date", "Time"]
    # Sheet 3: Hisib Errors
    he_headers = ["Error code", "Error Description"]

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

    # Hisi Disconnects sheet
    if "Hisi Disconnects" in wb.sheetnames:
        ws_hd = wb["Hisi Disconnects"]
    else:
        ws_hd = wb.create_sheet("Hisi Disconnects")
        ws_hd.append(hd_headers)
    for d in hisi_disconnects:
        ws_hd.append([
            d.get("Date", ""),
            d.get("Time", ""),
            d.get("Last Sample No. before disconnect", ""),
            d.get("First Sample No. after reconnect", "")
        ])

    # Client Shutdowns sheet
    if "Client Shutdowns" in wb.sheetnames:
        ws_cs = wb["Client Shutdowns"]
    else:
        ws_cs = wb.create_sheet("Client Shutdowns")
        ws_cs.append(cs_headers)
    for d in client_shutdowns:
        date = d.get("Date", "")
        if date and len(date) == 6 and date.isdigit():
            date = datetime.strptime(date, "%y%m%d").strftime("%d-%b-%y")
        ws_cs.append([
            date,
            d.get("Time", "")
        ])

    # Hisib Errors sheet
    if "Hisib Errors" in wb.sheetnames:
        ws_he = wb["Hisib Errors"]
    else:
        ws_he = wb.create_sheet("Hisib Errors")
        ws_he.append(he_headers)
    for d in hisib_errors:
        ws_he.append([
            d.get("Error code", ""),
            d.get("Error name", "")
        ])

    wb.save(excel_path)


def extract_sample_numbers(log_path, excel_path="output.xlsx"):
    with open(log_path, 'r') as file:
        lines = file.readlines()
    hisi_disconnects = []
    client_shutdowns = []
    hisib_errors = []
    occurrence = 0
    client_shutdown_occurrence = 0
    i = 0
    while i < len(lines):
        line = lines[i]
        if "Detected HISIB Disconnect" in line:
            occurrence += 1
            # Improved regex: match [D] or [I] with optional whitespace, date and time
            match = re.search(r"\[.\]\s*(\d{6}) (\d{2}:\d{2}:\d{2})", line)
            date = match.group(1) if match else ""
            time = match.group(2) if match else ""
            # Format date for consistency
            if date and len(date) == 6 and date.isdigit():
                date_fmt = datetime.strptime(date, "%y%m%d").strftime("%d-%b-%y")
            else:
                date_fmt = date
            last_sample_no = None
            first_sample_no = None
            lookahead = lines[i+1:i+6]
            for l in lookahead:
                if "Last SampleNo before disconnect" in l:
                    parts = l.split('=')
                    if len(parts) > 1:
                        last_sample_no = parts[1].strip()
                if "First SampleNo after reconnect" in l:
                    parts = l.split('=')
                    if len(parts) > 1:
                        first_sample_no = parts[1].strip()
            hisi_disconnects.append({
                "Date": date_fmt,
                "Time": time,
                "Last Sample No. before disconnect": last_sample_no,
                "First Sample No. after reconnect": first_sample_no
            })
            i += 5
        elif "Command channel: Client shutdown" in line:
            match = re.search(r'(\d{6}) (\d{2}:\d{2}:\d{2}) .*Client shutdown', line)
            if match:
                client_shutdown_occurrence += 1
                date = match.group(1)
                time = match.group(2)
                client_shutdowns.append({
                    "Date": date,
                    "Time": time
                })
            i += 1
        else:
            # Check for setHisibErrorStatus = 0x400000XX
            err_match = re.search(r'setHisibErrorStatus\s*=\s*0x[0-9A-Fa-f]{6}([0-9A-Fa-f]{2})', line)
            if err_match:
                err_code = err_match.group(1).upper()
                err_name = ERROR_TABLE.get(err_code, f'Unknown error code {err_code}')
                hisib_errors.append({
                    "Error code": err_code,
                    "Error name": err_name
                })
            i += 1
    output = {
        "Hisi Disconnects": hisi_disconnects,
        "Client Shutdowns": client_shutdowns,
        "Hisib Errors": hisib_errors
    }
    print(json.dumps(output, indent=2))

    # Excel output (in table format as per user screenshot)
    append_events_to_excel(hisi_disconnects, client_shutdowns, hisib_errors, excel_path)


def run_gui():
    def browse_log():
        path = filedialog.askopenfilename(title="Select Log File", filetypes=[("Log Files", "*.log"), ("All Files", "*")])
        if path:
            log_entry.delete(0, tk.END)
            log_entry.insert(0, path)

    def browse_excel():
        path = filedialog.asksaveasfilename(title="Save Excel File As", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if path:
            excel_entry.delete(0, tk.END)
            excel_entry.insert(0, path)

    def run_extraction():
        log_path = log_entry.get()
        excel_path = excel_entry.get()
        if not log_path or not excel_path:
            messagebox.showerror("Error", "Please provide both input log file and output Excel file.")
            return
        try:
            extract_sample_numbers(log_path, excel_path)
            messagebox.showinfo("Success", f"Extraction complete!\nOutput saved to:\n{excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

    root = tk.Tk()
    root.title("Log to Excel Extractor")
    root.geometry("480x180")

    tk.Label(root, text="Input Log File:").pack(anchor="w", padx=10, pady=(10,0))
    log_frame = tk.Frame(root)
    log_frame.pack(fill="x", padx=10)
    log_entry = tk.Entry(log_frame, width=50)
    log_entry.pack(side="left", fill="x", expand=True)
    tk.Button(log_frame, text="Browse", command=browse_log).pack(side="left", padx=5)

    tk.Label(root, text="Output Excel File:").pack(anchor="w", padx=10, pady=(10,0))
    excel_frame = tk.Frame(root)
    excel_frame.pack(fill="x", padx=10)
    excel_entry = tk.Entry(excel_frame, width=50)
    excel_entry.pack(side="left", fill="x", expand=True)
    tk.Button(excel_frame, text="Browse", command=browse_excel).pack(side="left", padx=5)

    tk.Button(root, text="Run", command=run_extraction, height=2, width=10).pack(pady=15)
    root.mainloop()


if __name__ == "__main__":
    run_gui()
